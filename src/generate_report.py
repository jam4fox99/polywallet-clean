"""
Generate Excel report from Supabase cached data - ZERO API CALLS.
Matches exact format of report.py output.
"""
import sys
sys.path.insert(0, ".")
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from src import db_cache

OUTPUT_DIR = Path(__file__).parent.parent / "output"

# Styles
GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
CAT_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
LABEL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), 
                top=Side(style='thin'), bottom=Side(style='thin'))

def pnl_fill(val):
    return GREEN if val >= 0 else RED

def add_wallet_sheet(wb, wallet, stats, price_tiers, categories, positions):
    """Add a wallet sheet matching original report.py format."""
    ws = wb.create_sheet(title=(stats.get("username") or wallet[:20])[:31])
    
    # Row 1-3: Wallet info
    ws.cell(row=1, column=1, value="WALLET").font = Font(bold=True)
    ws.cell(row=1, column=1).fill = LABEL
    ws.cell(row=1, column=2, value=wallet)
    
    ws.cell(row=2, column=1, value="Username").fill = LABEL
    ws.cell(row=2, column=2, value=stats.get("username", ""))
    ws.cell(row=2, column=3, value="Rank").fill = LABEL
    ws.cell(row=2, column=4, value=f"#{stats.get('rank', '')}")
    
    # Row 4-10: PNL Breakdown
    ws["A4"] = "PNL BREAKDOWN"
    ws["A4"].font = Font(bold=True, size=10)
    
    for col, h in enumerate(["Period", "Realized", "Unrealized", "Total"], 1):
        c = ws.cell(row=5, column=col, value=h)
        c.fill = HEADER
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.border = BORDER
    
    realized_1d = float(stats.get("realized_1d") or 0)
    realized_7d = float(stats.get("realized_7d") or 0)
    realized_30d = float(stats.get("realized_30d") or 0)
    realized_all = float(stats.get("realized_all") or 0)
    unrealized = float(stats.get("unrealized_pnl") or 0)
    
    pnl_rows = [
        ("1D", realized_1d, None, realized_1d),
        ("7D", realized_7d, None, realized_7d),
        ("30D", realized_30d, None, realized_30d),
        ("All", realized_all, unrealized, realized_all + unrealized),
    ]
    
    for row_idx, (label, realized, unreal, total) in enumerate(pnl_rows, 6):
        ws.cell(row=row_idx, column=1, value=label).border = BORDER
        
        c = ws.cell(row=row_idx, column=2, value=realized)
        c.number_format = '"$"#,##0.00'
        c.fill = pnl_fill(realized)
        c.border = BORDER
        
        if unreal is None:
            ws.cell(row=row_idx, column=3, value="-").border = BORDER
        else:
            c = ws.cell(row=row_idx, column=3, value=unreal)
            c.number_format = '"$"#,##0.00'
            c.fill = pnl_fill(unreal)
            c.border = BORDER
        
        c = ws.cell(row=row_idx, column=4, value=total)
        c.number_format = '"$"#,##0.00'
        c.fill = pnl_fill(total)
        c.border = BORDER
    
    # Row 11-17: Stats
    ws["A11"] = "STATS"
    ws["A11"].font = Font(bold=True, size=10)
    
    volume = float(stats.get("volume") or 0)
    roi = float(stats.get("roi") or 0)
    win_rate = float(stats.get("win_rate") or 0)
    wins = int(stats.get("wins") or 0)
    losses = int(stats.get("losses") or 0)
    markets = int(stats.get("markets_traded") or 0)
    trades = int(stats.get("total_trades") or 0)
    avg_bet = float(stats.get("avg_bet_size") or 0)
    lb_pnl = float(stats.get("lb_pnl") or 0)
    calc_pnl = float(stats.get("calc_pnl") or 0)
    
    stats_data = [
        (12, "Volume", volume, "ROI", f"{roi}%"),
        (13, "Win Rate", f"{win_rate}%", "Wins", wins),
        (14, "Losses", losses, "Markets", markets),
        (15, "Trades", trades, "Avg Bet Size", avg_bet),
        (16, "LB PnL", lb_pnl, "Calc PnL", calc_pnl),
    ]
    
    for row, lbl1, val1, lbl2, val2 in stats_data:
        ws.cell(row=row, column=1, value=lbl1).fill = LABEL
        c1 = ws.cell(row=row, column=2, value=val1)
        if isinstance(val1, (int, float)) and lbl1 in ["Volume", "LB PnL", "Calc PnL", "Avg Bet Size"]:
            c1.number_format = '"$"#,##0.00'
        ws.cell(row=row, column=3, value=lbl2).fill = LABEL
        c2 = ws.cell(row=row, column=4, value=val2)
        if isinstance(val2, (int, float)) and lbl2 in ["LB PnL", "Calc PnL", "Avg Bet Size"]:
            c2.number_format = '"$"#,##0.00'
    
    # Col F-H: Categories Summary
    ws["F1"] = "CATEGORIES SUMMARY"
    ws["F1"].font = Font(bold=True, size=10)
    
    for col, h in enumerate(["Category", "% Volume", "PnL"], 6):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = HEADER
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.border = BORDER
    
    for i, cat in enumerate(categories[:8], 3):
        ws.cell(row=i, column=6, value=cat.get("category", "")).border = BORDER
        ws.cell(row=i, column=7, value=f"{cat.get('pct_volume', 0)}%").border = BORDER
        pnl = float(cat.get("pnl") or 0)
        c = ws.cell(row=i, column=8, value=pnl)
        c.number_format = '"$"#,##0.00'
        c.fill = pnl_fill(pnl)
        c.border = BORDER
    
    # Row 18-28: Price Tiers
    ws["A18"] = "PRICE TIERS"
    ws["A18"].font = Font(bold=True, size=10)
    
    for col, h in enumerate(["Tier", "Pos", "%", "Win%", "PnL"], 1):
        c = ws.cell(row=19, column=col, value=h)
        c.fill = HEADER
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.border = BORDER
    
    for i, tier in enumerate(price_tiers, 20):
        ws.cell(row=i, column=1, value=tier.get("tier", "")).border = BORDER
        ws.cell(row=i, column=2, value=tier.get("positions", 0)).border = BORDER
        ws.cell(row=i, column=3, value=f"{tier.get('pct_of_total', 0)}%").border = BORDER
        ws.cell(row=i, column=4, value=f"{tier.get('win_rate', 0)}%").border = BORDER
        pnl = float(tier.get("total_pnl") or 0)
        c = ws.cell(row=i, column=5, value=pnl)
        c.number_format = '"$"#,##0.00'
        c.fill = pnl_fill(pnl)
        c.border = BORDER
    
    # Row 32+: Positions grouped by category
    cat_positions = defaultdict(list)
    for pos in positions:
        cat_positions[pos.get("category", "Other")].append(pos)
    
    cat_order = sorted(cat_positions.keys(), 
                      key=lambda c: sum(float(p.get("pnl") or 0) for p in cat_positions[c]), 
                      reverse=True)
    
    COLS_PER_CAT = 7
    CAT_START_ROW = 32
    
    for cat_idx, cat_name in enumerate(cat_order):
        cat_trades = cat_positions[cat_name]
        cat_pnl = sum(float(p.get("pnl") or 0) for p in cat_trades)
        start_col = 1 + (cat_idx * COLS_PER_CAT)
        
        # Category header
        header_cell = ws.cell(row=CAT_START_ROW, column=start_col, 
                              value=f"{cat_name} ({len(cat_trades)})")
        header_cell.font = Font(bold=True, color="FFFFFF", size=10)
        header_cell.fill = CAT_HEADER
        
        # Total PnL
        pnl_cell = ws.cell(row=CAT_START_ROW + 1, column=start_col, 
                           value=f"Total PnL: ${cat_pnl:,.2f}")
        pnl_cell.font = Font(bold=True, size=9)
        pnl_cell.fill = pnl_fill(cat_pnl)
        
        # Column headers
        col_headers = ["Market", "Out", "Amt", "PnL", "ROI%", "Entry", "Exit"]
        for h_idx, h in enumerate(col_headers):
            c = ws.cell(row=CAT_START_ROW + 2, column=start_col + h_idx, value=h)
            c.fill = HEADER
            c.font = Font(bold=True, color="FFFFFF", size=8)
            c.border = BORDER
        
        # Position rows
        for t_idx, pos in enumerate(cat_trades):
            row = CAT_START_ROW + 3 + t_idx
            
            market_name = pos.get("market_name", "")[:25] if pos.get("market_name") else ""
            ws.cell(row=row, column=start_col, value=market_name).border = BORDER
            ws.cell(row=row, column=start_col + 1, value=(pos.get("outcome", "")[:8])).border = BORDER
            
            amt = float(pos.get("usd_amount") or 0)
            c = ws.cell(row=row, column=start_col + 2, value=amt)
            c.number_format = '"$"#,##0'
            c.border = BORDER
            
            pnl = float(pos.get("pnl") or 0)
            c = ws.cell(row=row, column=start_col + 3, value=pnl)
            c.number_format = '"$"#,##0.00'
            c.fill = pnl_fill(pnl)
            c.border = BORDER
            
            roi = float(pos.get("roi") or 0)
            c = ws.cell(row=row, column=start_col + 4, value=f"{roi}%")
            c.fill = pnl_fill(roi)
            c.border = BORDER
            
            entry_price = float(pos.get("entry_price") or 0)
            ws.cell(row=row, column=start_col + 5, value=f"${entry_price:.2f}").border = BORDER
            ws.cell(row=row, column=start_col + 6, value=pos.get("exit_date", "")).border = BORDER
    
    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    
    col_widths = [26, 8, 12, 14, 10, 12, 12]
    for cat_idx in range(len(cat_order)):
        for w_idx, width in enumerate(col_widths):
            col_num = (cat_idx * COLS_PER_CAT) + w_idx + 1
            if col_num <= 702:
                ws.column_dimensions[get_column_letter(col_num)].width = width

def generate_report(wallets=None, limit=30, output_file=None):
    """Generate full report for wallets from cached data."""
    client = db_cache.get_client()
    
    if wallets is None:
        result = client.table("wallet_stats").select("wallet").order("rank").limit(limit).execute()
        wallets = [r["wallet"] for r in result.data]
    
    print(f"Generating report for {len(wallets)} wallets...", flush=True)
    
    wb = Workbook()
    wb.remove(wb.active)
    
    for wallet in wallets:
        # Fetch all data from Supabase
        stats_result = client.table("wallet_stats").select("*").eq("wallet", wallet).execute()
        if not stats_result.data:
            print(f"  Skipping {wallet[:10]}... (no stats)", flush=True)
            continue
        stats = stats_result.data[0]
        
        tiers = client.table("wallet_price_tiers").select("*").eq("wallet", wallet).order("tier_order").execute()
        categories = client.table("wallet_categories").select("*").eq("wallet", wallet).order("pnl", desc=True).execute()
        positions = client.table("positions_enriched").select("*").eq("wallet", wallet).order("pnl", desc=True).execute()
        
        add_wallet_sheet(wb, wallet, stats, tiers.data, categories.data, positions.data)
        print(f"  Added: {stats.get('username') or wallet[:20]}", flush=True)
    
    if output_file is None:
        output_file = OUTPUT_DIR / "weekly_report.xlsx"
    
    wb.save(output_file)
    print(f"\nSaved: {output_file}", flush=True)
    return output_file

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=30, help="Number of wallets")
    parser.add_argument("--output", help="Output file path")
    args = parser.parse_args()
    
    generate_report(limit=args.limit, output_file=args.output)
