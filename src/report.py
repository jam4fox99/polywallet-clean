"""
DIAGNOSTIC VERSION - Very detailed logging to find the hang
Now with Supabase caching for incremental syncing!
"""
import os
import argparse
import asyncio
from pathlib import Path
import aiohttp
import pandas as pd
import time
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import traceback
from src import db_cache

BASE_URL = "https://data-api.polymarket.com"
GAMMA_URL = "https://gamma-api.polymarket.com"
PROXY_URL = os.getenv("PROXY_URL")
ROOT_DIR = Path(__file__).resolve().parent.parent  # Go up from src/ to project root
DEFAULT_WALLET_FILE = ROOT_DIR / "data" / "wallets.csv"
DEFAULT_OUTPUT_FILE = ROOT_DIR / "output" / "report.xlsx"

MAX_CONCURRENT_WALLETS = 30  # Keep moderate concurrency for stability
MAX_CONCURRENT_CATEGORIES = 10  # Keep category lookups gentle to avoid rate limits
REQUEST_TIMEOUT = 5
MAX_RETRIES = 2  # Reduced
THIRTY_DAYS_SEC = 30 * 86400
WALLET_TIMEOUT = None  # No per-wallet timeout (run until all pages fetched)
PAGES_PER_BATCH = 20  # pages to fetch in parallel per wallet/endpoint

PRICE_TIERS = [(90,100,"90-100c"),(80,90,"80-90c"),(70,80,"70-80c"),(60,70,"60-70c"),(50,60,"50-60c"),(40,50,"40-50c"),(30,40,"30-40c"),(20,30,"20-30c"),(10,20,"10-20c"),(0,10,"0-10c")]

class MultiWalletReportGenerator:
    def __init__(self, proxy_url: str | None = PROXY_URL, use_cache: bool = True):
        self.api_calls = 0
        self.retries = 0
        self.errors = 0
        self.skipped = 0
        self.proxy_url = proxy_url
        self.use_cache = use_cache and db_cache.is_cache_enabled()
        self.market_tags_cache = {}
        self.cache_lock = asyncio.Lock()
        self.wallet_semaphore = asyncio.Semaphore(MAX_CONCURRENT_WALLETS)
        self.completed = 0
        self.total = 0
        self.start_time = None
        self.active_wallets = {}
        self.results_list = []  # Store results as we go
        self.cache_hits = 0
        self.new_trades_fetched = 0
        
    async def fetch(self, session, url, params=None):
        for attempt in range(MAX_RETRIES):
            self.api_calls += 1
            try:
                async with session.get(url, params=params, proxy=self.proxy_url,
                                       timeout=aiohttp.ClientTimeout(total=REQUEST_TIMEOUT)) as r:
                    if r.status == 200:
                        return await r.json()
                    self.retries += 1
            except Exception as e:
                self.retries += 1
        self.errors += 1
        return None

    async def fetch_all_paginated(self, session, endpoint, wallet, limit=100, extra_params=None, wallet_short="", batch_size=PAGES_PER_BATCH, max_pages=None):
        """Fetch pages in parallel batches until an empty/short page is returned. Optional max_pages cap."""
        all_data = []
        offset = 0
        page = 0
        while True:
            if max_pages:
                remaining = max_pages - page
                if remaining <= 0:
                    break
                current_batch = min(batch_size, remaining)
            else:
                current_batch = batch_size

            tasks = []
            for i in range(current_batch):
                self.active_wallets[wallet_short] = f"{endpoint} pg{page + i + 1}"
                params = {"user": wallet, "limit": limit, "offset": offset + i * limit}
                if extra_params:
                    params.update(extra_params)
                tasks.append(self.fetch(session, f"{BASE_URL}/{endpoint}", params))

            pages = await asyncio.gather(*tasks)
            any_short = False
            added = 0
            for p in pages:
                if p:
                    all_data.extend(p)
                    added += len(p)
                    if len(p) < limit:
                        any_short = True
                else:
                    any_short = True

            offset += limit * current_batch
            page += current_batch

            if any_short or added == 0:
                break

        return all_data

    async def get_market_tags(self, session, slug, category_sem):
        if not slug:
            return slug, []
        async with self.cache_lock:
            if slug in self.market_tags_cache:
                return slug, self.market_tags_cache[slug]
        
        # Check Supabase cache
        if self.use_cache:
            db_tags = db_cache.get_cached_market_tags([slug])
            if slug in db_tags:
                async with self.cache_lock:
                    self.market_tags_cache[slug] = db_tags[slug]
                return slug, db_tags[slug]
        
        tags = []
        async with category_sem:
            data = await self.fetch(session, f"{GAMMA_URL}/markets", {"slug": slug})
            if data and len(data) > 0:
                market_id = data[0].get("id")
                if market_id:
                    tag_data = await self.fetch(session, f"{GAMMA_URL}/markets/{market_id}/tags")
                    if tag_data:
                        tags = [t.get("label") for t in tag_data if t.get("label") and t.get("label") != "All"]
        
        async with self.cache_lock:
            self.market_tags_cache[slug] = tags
        
        # Save to Supabase
        if self.use_cache and tags:
            db_cache.save_market_tags(slug, tags)
        
        return slug, tags

    async def fetch_wallet_data(self, session, wallet):
        ws = wallet[:10]
        # Fetch leaderboard for all time periods
        self.active_wallets[ws] = "leaderboard"
        lb_all = await self.fetch(session, f"{BASE_URL}/v1/leaderboard", {"timePeriod": "all", "user": wallet})
        lb_day = await self.fetch(session, f"{BASE_URL}/v1/leaderboard", {"timePeriod": "day", "user": wallet})
        lb_week = await self.fetch(session, f"{BASE_URL}/v1/leaderboard", {"timePeriod": "week", "user": wallet})
        lb_month = await self.fetch(session, f"{BASE_URL}/v1/leaderboard", {"timePeriod": "month", "user": wallet})
        lb = lb_all  # Use all-time for main stats
        lb_periods = {"all": lb_all, "day": lb_day, "week": lb_week, "month": lb_month}
        self.active_wallets[ws] = "traded"
        traded = await self.fetch(session, f"{BASE_URL}/traded", {"user": wallet})
        
        # Check cache for existing trades
        cached_trades = []
        last_ts = 0
        if self.use_cache:
            self.active_wallets[ws] = "cache-check"
            last_ts = db_cache.get_last_trade_timestamp(wallet)
            if last_ts > 0:
                cached_trades = db_cache.get_cached_trades(wallet)
                self.cache_hits += len(cached_trades)
                print(f"  [Cache] {ws}: Found {len(cached_trades)} cached trades, last_ts={last_ts}")
        
        # Fetch trades - if we have cached data, fetch incrementally
        self.active_wallets[ws] = "trades"
        if last_ts > 0 and cached_trades:
            # Fetch new trades only (pages until we hit old timestamps)
            new_trades = await self.fetch_trades_incremental(session, wallet, last_ts, ws)
            trades = new_trades + cached_trades
            self.new_trades_fetched += len(new_trades)
            print(f"  [Cache] {ws}: Fetched {len(new_trades)} new trades, total={len(trades)}")
        else:
            # First time - fetch all
            trades = await self.fetch_all_paginated(session, "trades", wallet, 500, {}, ws, max_pages=5000)
            self.new_trades_fetched += len(trades)
        
        self.active_wallets[ws] = "closed-pos"
        closed = await self.fetch_all_paginated(session, "closed-positions", wallet, 50, {"sortBy": "realizedpnl", "sortDirection": "DESC"}, ws)
        self.active_wallets[ws] = "positions"
        positions = await self.fetch_all_paginated(session, "positions", wallet, 500, {"sortBy": "CURRENT", "sortDirection": "DESC"}, ws)
        
        # Save to cache
        if self.use_cache:
            self.active_wallets[ws] = "cache-save"
            username = lb[0].get("userName", "") if lb and len(lb) > 0 else ""
            rank = lb[0].get("rank") if lb and len(lb) > 0 else None
            db_cache.save_wallet(wallet, username, rank)
            # Only save new trades (not cached ones)
            if trades:
                trades_to_save = [t for t in trades if t.get("timestamp", 0) > last_ts] if last_ts > 0 else trades
                if trades_to_save:
                    db_cache.save_trades(wallet, trades_to_save)
            if closed:
                db_cache.save_closed_positions(wallet, closed)
            # Save open positions (always fresh)
            db_cache.save_open_positions(wallet, positions or [])
            # Save leaderboard rankings for all time periods
            db_cache.save_wallet_leaderboard_stats(wallet, lb_periods)
        
        return {"leaderboard": lb, "traded": traded, "trades": trades, "closed": closed, "positions": positions, "lb_periods": lb_periods}
    
    async def fetch_trades_incremental(self, session, wallet, last_ts, wallet_short):
        """Fetch only new trades after last_ts timestamp."""
        all_new = []
        offset = 0
        limit = 500
        page = 0
        
        while True:
            self.active_wallets[wallet_short] = f"trades-inc pg{page + 1}"
            params = {"user": wallet, "limit": limit, "offset": offset}
            data = await self.fetch(session, f"{BASE_URL}/trades", params)
            
            if not data:
                break
            
            # Filter to only trades newer than last_ts
            new_in_page = [t for t in data if t.get("timestamp", 0) > last_ts]
            all_new.extend(new_in_page)
            
            # If we found older trades, we've caught up - stop
            if len(new_in_page) < len(data):
                break
            
            # If page was short, we've reached the end
            if len(data) < limit:
                break
            
            offset += limit
            page += 1
            
            # Safety limit
            if page > 100:
                break
        
        return all_new

    def calculate_stats(self, data):
        stats = {}
        lb = data["leaderboard"]
        if lb and len(lb) > 0:
            stats["total_pnl"] = float(lb[0].get("pnl", 0))
            stats["volume"] = float(lb[0].get("vol", 0))
            stats["rank"] = lb[0].get("rank", "N/A")
            stats["username"] = lb[0].get("userName", "")
        else:
            stats["total_pnl"], stats["volume"], stats["rank"], stats["username"] = 0, 0, "N/A", ""
        
        stats["markets_traded"] = data["traded"].get("traded", 0) if data["traded"] else 0
        trades, closed, positions = data["trades"] or [], data["closed"] or [], data["positions"] or []
        stats["total_trades"], stats["closed_positions"], stats["open_positions"] = len(trades), len(closed), len(positions)
        
        # Fix: Include realizedPnl from BOTH closed AND open positions
        realized_closed = sum(float(p.get("realizedPnl", 0)) for p in closed)
        realized_open = sum(float(p.get("realizedPnl", 0)) for p in positions)
        stats["realized_pnl"] = realized_closed + realized_open
        stats["unrealized_pnl"] = sum(float(p.get("cashPnl", 0)) for p in positions)
        
        # Avg bet size: total volume / number of positions
        total_positions = len(closed) + len(positions)
        stats["avg_bet_size"] = round(stats["volume"] / total_positions, 2) if total_positions > 0 else 0
        
        now = time.time()
        stats["realized_1d"] = sum(float(p.get("realizedPnl", 0)) for p in closed if p.get("timestamp", 0) > now - 86400)
        stats["realized_7d"] = sum(float(p.get("realizedPnl", 0)) for p in closed if p.get("timestamp", 0) > now - 604800)
        stats["realized_30d"] = sum(float(p.get("realizedPnl", 0)) for p in closed if p.get("timestamp", 0) > now - 2592000)
        stats["unrealized_1d"] = stats["unrealized_7d"] = stats["unrealized_30d"] = stats["unrealized_pnl"]
        
        wins = sum(1 for p in closed if float(p.get("realizedPnl", 0)) > 0)
        losses = sum(1 for p in closed if float(p.get("realizedPnl", 0)) < 0)
        stats["wins"], stats["losses"] = wins, losses
        stats["win_rate"] = round((wins / (wins + losses)) * 100, 2) if (wins + losses) > 0 else 0
        
        if trades:
            stats["avg_trade_size"] = round(sum(float(t.get("size", 0)) * float(t.get("price", 0)) for t in trades) / len(trades), 2)
            ts = [t.get("timestamp", 0) for t in trades if t.get("timestamp")]
            if ts:
                days = max((max(ts) - min(ts)) / 86400, 1)
                stats["days_active"], stats["trades_per_day"] = round(days, 1), round(len(trades) / days, 2)
            else:
                stats["days_active"], stats["trades_per_day"] = 0, 0
        else:
            stats["avg_trade_size"], stats["days_active"], stats["trades_per_day"] = 0, 0, 0
        
        stats["roi"] = round((stats["total_pnl"] / stats["volume"]) * 100, 2) if stats["volume"] > 0 else 0
        stats["calc_total_pnl"] = stats["realized_pnl"] + stats["unrealized_pnl"]
        stats["pnl_diff"] = stats["total_pnl"] - stats["calc_total_pnl"]
        return stats

    def calculate_price_tiers(self, closed):
        tiers = {t[2]: {"positions": 0, "wins": 0, "losses": 0, "total_pnl": 0.0} for t in PRICE_TIERS}
        for pos in closed:
            pnl = float(pos.get("realizedPnl", 0))
            entry_cents = int(float(pos.get("avgPrice", 0)) * 100)
            tier_name = next((name for low, high, name in PRICE_TIERS if low <= entry_cents < high), "90-100c" if entry_cents >= 100 else "0-10c")
            tiers[tier_name]["positions"] += 1
            tiers[tier_name]["total_pnl"] += pnl
            if pnl > 0: tiers[tier_name]["wins"] += 1
            elif pnl < 0: tiers[tier_name]["losses"] += 1
        
        total_pos = sum(t["positions"] for t in tiers.values())
        return [{"tier": name, "positions": tiers[name]["positions"],
                 "pct_of_total": round((tiers[name]["positions"] / total_pos) * 100, 1) if total_pos > 0 else 0,
                 "win_rate": round((tiers[name]["wins"] / (tiers[name]["wins"] + tiers[name]["losses"])) * 100, 1) if (tiers[name]["wins"] + tiers[name]["losses"]) > 0 else 0,
                 "total_pnl": round(tiers[name]["total_pnl"], 2)} for _, _, name in PRICE_TIERS]

    async def fetch_categories_limited(self, session, closed, wallet_short, max_cats=None):
        """Fetch categories for ALL unique slugs (no slicing) with caching and gentle concurrency."""
        category_sem = asyncio.Semaphore(MAX_CONCURRENT_CATEGORIES)
        slugs = list(set(p.get("slug", "") for p in closed if p.get("slug")))

        async with self.cache_lock:
            uncached = [s for s in slugs if s not in self.market_tags_cache]

        if uncached:
            self.active_wallets[wallet_short] = f"cats 0/{len(uncached)}"
            await asyncio.gather(*[self.get_market_tags(session, slug, category_sem) for slug in uncached])

        cat_data = defaultdict(lambda: {"volume": 0.0, "pnl": 0.0})
        for pos in closed:
            tags = self.market_tags_cache.get(pos.get("slug", ""), [])
            cat = tags[0] if tags else "Unknown"
            cat_data[cat]["volume"] += float(pos.get("totalBought", 0)) * float(pos.get("avgPrice", 0))
            cat_data[cat]["pnl"] += float(pos.get("realizedPnl", 0))

        total_vol = sum(d["volume"] for d in cat_data.values())
        return [{"category": cat, "pct_volume": round((d["volume"] / total_vol) * 100, 1) if total_vol > 0 else 0,
                 "pnl": round(d["pnl"], 2)} for cat, d in sorted(cat_data.items(), key=lambda x: -x[1]["volume"])]

    def calculate_hold_times(self, trades):
        # Simplified - just return defaults if too many trades
        if len(trades) > 1000:
            return {"avg_minutes": 0, "avg_hours": 0, "min_minutes": 0}
        by_cond = defaultdict(list)
        for t in trades:
            if t.get("conditionId"):
                by_cond[t["conditionId"]].append(t)
        hold_times = []
        for ct in list(by_cond.values())[:100]:  # Limit
            buys = sorted([t for t in ct if t.get("side") == "BUY"], key=lambda x: x.get("timestamp", 0))
            sells = sorted([t for t in ct if t.get("side") == "SELL"], key=lambda x: x.get("timestamp", 0))
            for buy in buys[:10]:
                for sell in sells[:10]:
                    if sell.get("timestamp", 0) > buy.get("timestamp", 0):
                        hold_times.append((sell["timestamp"] - buy["timestamp"]) / 60)
                        break
        return {"avg_minutes": round(sum(hold_times) / len(hold_times), 1) if hold_times else 0,
                "avg_hours": round(sum(hold_times) / len(hold_times) / 60, 2) if hold_times else 0,
                "min_minutes": round(min(hold_times), 1) if hold_times else 0}

    def format_positions_with_dates(self, closed, trades):
        """Format ALL positions with dates, ROI, and category - NO LIMITS"""
        entry_dates = {}
        for t in trades:  # No limit - process all trades
            cid = t.get("conditionId", "")
            if cid and t.get("side") == "BUY":
                ts = t.get("timestamp", 0)
                if cid not in entry_dates or ts < entry_dates[cid]:
                    entry_dates[cid] = ts
        
        positions = []
        for p in closed:  # No limit - process ALL closed positions
            ts = p.get("timestamp", 0)
            # REMOVED: 30-day filter - show ALL trades
            cid = p.get("conditionId", "")
            entry_ts = entry_dates.get(cid, 0)
            entry_date = datetime.fromtimestamp(entry_ts).strftime("%Y-%m-%d") if entry_ts else ""
            exit_date = p.get("endDate", "")[:10] if p.get("endDate") else ""
            if not exit_date and ts:
                exit_date = datetime.fromtimestamp(ts).strftime("%Y-%m-%d")
            tags = self.market_tags_cache.get(p.get("slug", ""), [])
            
            usd_amount = round(float(p.get("avgPrice", 0)) * float(p.get("totalBought", 0)), 2)
            pnl = round(float(p.get("realizedPnl", 0)), 2)
            roi = round((pnl / usd_amount) * 100, 1) if usd_amount > 0 else 0
            
            positions.append({
                "market": p.get("title", "")[:40],
                "outcome": p.get("outcome", "")[:15],
                "entry_price": round(float(p.get("avgPrice", 0)), 3),
                "usd_amount": usd_amount,
                "pnl": pnl,
                "roi": roi,
                "entry_date": entry_date,
                "exit_date": exit_date,
                "category": (tags[0] if tags else "Other")
            })
        return positions

    def add_wallet_sheet(self, wb, wallet, stats, price_tiers, categories, hold_times, positions):
        ws = wb.create_sheet(title=wallet[:31])
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cat_header = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        label_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Row 1-3: Wallet info - SEPARATE CELLS
        ws.cell(row=1, column=1, value="WALLET").font = Font(bold=True)
        ws.cell(row=1, column=1).fill = label_fill
        ws.cell(row=1, column=2, value=wallet)
        
        ws.cell(row=2, column=1, value="Username").fill = label_fill
        ws.cell(row=2, column=2, value=stats['username'])
        ws.cell(row=2, column=3, value="Rank").fill = label_fill
        ws.cell(row=2, column=4, value=stats['rank'])
        
        # Row 4-10: PNL Breakdown
        ws["A4"], ws["A4"].font = "PNL BREAKDOWN", Font(bold=True, size=10)
        for col, h in enumerate(["Period", "Realized", "Unrealized", "Total"], 1):
            c = ws.cell(row=5, column=col, value=h)
            c.fill, c.font, c.border = header, Font(bold=True, color="FFFFFF", size=9), border
        
        # PNL rows - show "-" for unrealized in 1D/7D/30D since it's not historical data
        pnl_rows = [
            ("1D", stats['realized_1d'], None, stats['realized_1d']),  # Unrealized N/A
            ("7D", stats['realized_7d'], None, stats['realized_7d']),  # Unrealized N/A
            ("30D", stats['realized_30d'], None, stats['realized_30d']),  # Unrealized N/A
            ("All", stats['realized_pnl'], stats['unrealized_pnl'], stats['realized_pnl'] + stats['unrealized_pnl']),
        ]
        for row_idx, (label, realized, unrealized, total) in enumerate(pnl_rows, 6):
            ws.cell(row=row_idx, column=1, value=label).border = border
            # Realized
            c = ws.cell(row=row_idx, column=2, value=realized)
            c.number_format = '"$"#,##0.00'
            c.fill = green if realized >= 0 else red
            c.border = border
            # Unrealized - show "-" for time-bounded periods
            if unrealized is None:
                c = ws.cell(row=row_idx, column=3, value="-")
                c.border = border
            else:
                c = ws.cell(row=row_idx, column=3, value=unrealized)
                c.number_format = '"$"#,##0.00'
                c.fill = green if unrealized >= 0 else red
                c.border = border
            # Total
            c = ws.cell(row=row_idx, column=4, value=total)
            c.number_format = '"$"#,##0.00'
            c.fill = green if total >= 0 else red
            c.border = border
        
        # Row 11-17: Stats - SEPARATE CELLS
        ws["A11"], ws["A11"].font = "STATS", Font(bold=True, size=10)
        stats_data = [
            (12, "Volume", stats['volume'], "ROI", f"{stats['roi']}%"),
            (13, "Win Rate", f"{stats['win_rate']}%", "Wins", stats['wins']),
            (14, "Losses", stats['losses'], "Markets", stats['markets_traded']),
            (15, "Trades", stats['total_trades'], "Avg Bet Size", stats['avg_bet_size']),
            (16, "LB PnL", stats['total_pnl'], "Calc PnL", stats['calc_total_pnl']),
        ]
        for row, lbl1, val1, lbl2, val2 in stats_data:
            ws.cell(row=row, column=1, value=lbl1).fill = label_fill
            c1 = ws.cell(row=row, column=2, value=val1)
            if isinstance(val1, (int, float)) and lbl1 in ["Volume", "LB PnL", "Calc PnL", "Avg Bet Size"]:
                c1.number_format = '"$"#,##0.00'
            ws.cell(row=row, column=3, value=lbl2).fill = label_fill
            c2 = ws.cell(row=row, column=4, value=val2)
            if isinstance(val2, (int, float)) and lbl2 in ["LB PnL", "Calc PnL", "Avg Bet Size"]:
                c2.number_format = '"$"#,##0.00'
        
        # Row 1-10 Col F-H: Categories Summary - PROPER TABLE
        ws["F1"], ws["F1"].font = "CATEGORIES SUMMARY", Font(bold=True, size=10)
        for col, h in enumerate(["Category", "% Volume", "PnL"], 6):
            c = ws.cell(row=2, column=col, value=h)
            c.fill, c.font, c.border = header, Font(bold=True, color="FFFFFF", size=9), border
        for i, cat in enumerate(categories[:8], 3):
            ws.cell(row=i, column=6, value=cat['category']).border = border
            ws.cell(row=i, column=7, value=f"{cat['pct_volume']}%").border = border
            c = ws.cell(row=i, column=8, value=cat['pnl'])
            c.number_format = '"$"#,##0.00'
            c.fill = green if cat['pnl'] >= 0 else red
            c.border = border
        
        # Row 18-28: Price Tiers
        ws["A18"], ws["A18"].font = "PRICE TIERS", Font(bold=True, size=10)
        for col, h in enumerate(["Tier", "Pos", "%", "Win%", "PnL"], 1):
            c = ws.cell(row=19, column=col, value=h)
            c.fill, c.font, c.border = header, Font(bold=True, color="FFFFFF", size=9), border
        for i, tier in enumerate(price_tiers, 20):
            ws.cell(row=i, column=1, value=tier["tier"]).border = border
            ws.cell(row=i, column=2, value=tier["positions"]).border = border
            ws.cell(row=i, column=3, value=f"{tier['pct_of_total']}%").border = border
            ws.cell(row=i, column=4, value=f"{tier['win_rate']}%").border = border
            c = ws.cell(row=i, column=5, value=tier['total_pnl'])
            c.number_format = '"$"#,##0.00'
            c.fill = green if tier['total_pnl'] >= 0 else red
            c.border = border
        
        # Group positions by category
        cat_positions = defaultdict(list)
        for pos in positions:
            cat_positions[pos["category"]].append(pos)
        
        # Sort categories by total PnL (highest first)
        cat_order = sorted(cat_positions.keys(), 
                          key=lambda c: sum(p["pnl"] for p in cat_positions[c]), reverse=True)
        
        # Horizontal category sections - each category gets 7 columns
        COLS_PER_CAT = 7  # Market, Outcome, Amt, PnL, ROI%, Entry, Exit
        CAT_START_ROW = 32
        
        for cat_idx, cat_name in enumerate(cat_order):
            cat_trades = cat_positions[cat_name]
            cat_pnl = sum(p["pnl"] for p in cat_trades)
            start_col = 1 + (cat_idx * COLS_PER_CAT)
            
            # Category header with total PnL
            header_cell = ws.cell(row=CAT_START_ROW, column=start_col, 
                                  value=f"{cat_name} ({len(cat_trades)})")
            header_cell.font = Font(bold=True, size=10)
            header_cell.fill = cat_header
            header_cell.font = Font(bold=True, color="FFFFFF", size=10)
            
            # Total PnL for category
            pnl_cell = ws.cell(row=CAT_START_ROW + 1, column=start_col, 
                               value=f"Total PnL: ${cat_pnl:,.2f}")
            pnl_cell.font = Font(bold=True, size=9)
            pnl_cell.fill = green if cat_pnl >= 0 else red
            
            # Column headers
            col_headers = ["Market", "Out", "Amt", "PnL", "ROI%", "Entry", "Exit"]
            for h_idx, h in enumerate(col_headers):
                c = ws.cell(row=CAT_START_ROW + 2, column=start_col + h_idx, value=h)
                c.fill, c.font, c.border = header, Font(bold=True, color="FFFFFF", size=8), border
            
            # Trade rows - NO LIMIT, show all trades
            for t_idx, trade in enumerate(cat_trades):
                row = CAT_START_ROW + 3 + t_idx
                
                ws.cell(row=row, column=start_col, value=trade["market"][:25]).border = border
                ws.cell(row=row, column=start_col + 1, value=trade["outcome"][:8]).border = border
                
                amt_cell = ws.cell(row=row, column=start_col + 2, value=trade["usd_amount"])
                amt_cell.number_format = '"$"#,##0'
                amt_cell.border = border
                
                pnl_c = ws.cell(row=row, column=start_col + 3, value=trade["pnl"])
                pnl_c.number_format = '"$"#,##0.00'
                pnl_c.fill = green if trade["pnl"] >= 0 else red
                pnl_c.border = border
                
                roi_c = ws.cell(row=row, column=start_col + 4, value=f"{trade['roi']}%")
                roi_c.fill = green if trade["roi"] >= 0 else red
                roi_c.border = border
                
                ws.cell(row=row, column=start_col + 5, value=trade["entry_date"]).border = border
                ws.cell(row=row, column=start_col + 6, value=trade["exit_date"]).border = border
        
        # Set column widths for top summary section (prevent #### values)
        ws.column_dimensions['A'].width = 15   # Labels
        ws.column_dimensions['B'].width = 45   # Wallet address / values
        ws.column_dimensions['C'].width = 15   # Unrealized / labels
        ws.column_dimensions['D'].width = 15   # Total / values
        ws.column_dimensions['E'].width = 15   # Price Tier PnL
        ws.column_dimensions['F'].width = 16   # Category names
        ws.column_dimensions['G'].width = 12   # % Volume
        ws.column_dimensions['H'].width = 15   # Category PnL
        
        # Set column widths for all category trade columns (starting col 1)
        col_widths = [26, 8, 12, 14, 10, 12, 12]  # Per category: Market, Out, Amt, PnL, ROI%, Entry, Exit
        for cat_idx in range(len(cat_order)):
            for w_idx, width in enumerate(col_widths):
                col_num = (cat_idx * COLS_PER_CAT) + w_idx + 1
                if col_num <= 702:  # Excel max columns (ZZ)
                    col_letter = get_column_letter(col_num)
                    ws.column_dimensions[col_letter].width = width

    async def process_wallet(self, session, wallet, idx):
        """Process one wallet with timeout"""
        ws = wallet[:10]
        
        async with self.wallet_semaphore:
            self.active_wallets[ws] = "starting"
            start = time.time()
            deadline = None  # No deadline; run to completion
            
            try:
                # STEP 1: Fetch data (full pagination)
                self.active_wallets[ws] = "fetch_data"
                data = await self.fetch_wallet_data(session, wallet)
                
                trades_count = len(data["trades"] or [])
                closed_count = len(data["closed"] or [])
                
                # STEP 2: Calculate stats (fast, no timeout needed)
                self.active_wallets[ws] = "calc_stats"
                stats = self.calculate_stats(data)
                tiers = self.calculate_price_tiers(data["closed"] or [])
                
                # STEP 3: Fetch categories
                self.active_wallets[ws] = "categories"
                cats = await self.fetch_categories_limited(session, data["closed"] or [], ws)
                
                # STEP 4: Hold times and positions (fast)
                self.active_wallets[ws] = "final"
                hold = self.calculate_hold_times(data["trades"] or [])
                pos = self.format_positions_with_dates(data["closed"] or [], data["trades"] or [])
                
                # STEP 5: Save all calculated stats to cache
                if self.use_cache:
                    self.active_wallets[ws] = "save-stats"
                    db_cache.save_wallet_stats(wallet, stats)
                    db_cache.save_price_tiers(wallet, tiers)
                    db_cache.save_categories(wallet, cats)
                    db_cache.save_hold_times(wallet, hold)
                
                wallet_time = time.time() - start
                self.completed += 1
                
                if ws in self.active_wallets:
                    del self.active_wallets[ws]
                
                print(f"  [{self.completed}/{self.total}] OK {ws} | {trades_count}t {closed_count}c | {wallet_time:.1f}s")
                return (wallet, stats, tiers, cats, hold, pos)
                
            except asyncio.TimeoutError:
                self.completed += 1
                self.skipped += 1
                stage = self.active_wallets.get(ws, "unknown")
                if ws in self.active_wallets:
                    del self.active_wallets[ws]
                print(f"  [{self.completed}/{self.total}] TIMEOUT {ws} at {stage}")
                return None
                
            except Exception as e:
                self.completed += 1
                self.errors += 1
                if ws in self.active_wallets:
                    del self.active_wallets[ws]
                print(f"  [{self.completed}/{self.total}] ERROR {ws}: {type(e).__name__}: {e}")
                return None

    async def progress_reporter(self):
        """Report progress every 5 seconds"""
        while self.completed < self.total:
            await asyncio.sleep(5)
            active_list = list(self.active_wallets.items())[:10]
            active_str = ", ".join([f"{w}:{s}" for w, s in active_list])
            print(f"  >> Progress: {self.completed}/{self.total} | Active({len(self.active_wallets)}): {active_str}")

    async def generate_multi_report(self, wallets, output_path):
        print("=" * 80)
        print(f"DIAGNOSTIC RUN - {len(wallets)} wallets")
        print(f"Concurrent: {MAX_CONCURRENT_WALLETS} | Timeout: {WALLET_TIMEOUT}s")
        print("=" * 80)
        
        self.total = len(wallets)
        self.start_time = time.time()
        wb = Workbook()
        wb.remove(wb.active)
        
        connector = aiohttp.TCPConnector(limit=100, ttl_dns_cache=300)
        
        async with aiohttp.ClientSession(connector=connector) as session:
            # Start progress reporter
            progress_task = asyncio.create_task(self.progress_reporter())
            
            # Process wallets
            tasks = [self.process_wallet(session, w, i) for i, w in enumerate(wallets)]
            results = await asyncio.gather(*tasks, return_exceptions=True)
            
            progress_task.cancel()
        
        print("\nBuilding XLSX...")
        successful = 0
        for r in results:
            if r and not isinstance(r, Exception):
                try:
                    self.add_wallet_sheet(wb, *r)
                    successful += 1
                except Exception as e:
                    print(f"  Sheet error: {e}")
        
        wb.save(output_path)
        elapsed = time.time() - self.start_time
        print(f"\nDONE! {elapsed:.1f}s ({elapsed/60:.1f} min)")
        print(f"OK: {successful} | Skipped: {self.skipped} | Errors: {self.errors}")
        if self.use_cache:
            print(f"Cache: {self.cache_hits} cached trades used | {self.new_trades_fetched} new trades fetched")
        print(f"Output: {output_path}")

def parse_args():
    parser = argparse.ArgumentParser(description="Generate XLSX reports for Polymarket wallets (v5 proxy diagnostic)")
    parser.add_argument(
        "--wallet-file",
        type=Path,
        default=DEFAULT_WALLET_FILE,
        help=f"CSV with wallet addresses (default: {DEFAULT_WALLET_FILE})",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_FILE,
        help=f"Output XLSX path (default: {DEFAULT_OUTPUT_FILE})",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Limit number of wallets (for quick tests)",
    )
    parser.add_argument(
        "--no-proxy",
        action="store_true",
        help="Disable BrightData proxy and call APIs directly",
    )
    parser.add_argument(
        "--no-cache",
        action="store_true",
        help="Disable Supabase caching (fetch all data fresh)",
    )
    return parser.parse_args()

async def main():
    args = parse_args()

    wallet_file = args.wallet_file.expanduser()
    output_path = args.output.expanduser()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if not wallet_file.is_file():
        raise FileNotFoundError(f"Wallet CSV not found: {wallet_file}")

    df = pd.read_csv(wallet_file)
    wallets = df["wallet"].tolist()
    if args.limit:
        wallets = wallets[: args.limit]
    print(f"Loaded {len(wallets)} wallets from {wallet_file}")
    if args.limit:
        print(f"Using first {args.limit} wallets (test mode)")
    print(f"Writing report to {output_path}")
    print(f"Proxy: {'disabled' if args.no_proxy else 'enabled'}")
    
    # Cache status
    use_cache = not args.no_cache
    cache_available = db_cache.is_cache_enabled()
    if use_cache and cache_available:
        print(f"Cache: ENABLED (Supabase)")
    elif use_cache and not cache_available:
        print(f"Cache: NOT CONFIGURED (set SUPABASE_URL and SUPABASE_KEY in .env)")
    else:
        print(f"Cache: DISABLED (--no-cache flag)")
    
    generator = MultiWalletReportGenerator(
        proxy_url=None if args.no_proxy else PROXY_URL,
        use_cache=use_cache
    )
    await generator.generate_multi_report(wallets, output_path)

if __name__ == "__main__":
    asyncio.run(main())
